"""Excel report generation with multi-sheet workbooks."""

from __future__ import annotations

import logging

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, LineChart
from openpyxl.utils import get_column_letter

from config import Config
from db.database import Database
from db import models

logger = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
CURRENCY_FORMAT = '₹#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


async def generate_report(db: Database, start: str | None = None, end: str | None = None) -> Path:
    """Generate a multi-sheet Excel report and return the file path."""
    today = date.today()
    if not start:
        start = today.replace(day=1).isoformat()
    if not end:
        end = today.isoformat()

    wb = Workbook()

    # --- Sheet 1: Expenses ---
    ws_exp = wb.active
    ws_exp.title = "Expenses"
    headers = ["Date", "Amount", "Category", "Subcategory", "Description", "Source"]
    ws_exp.append(headers)
    _style_header(ws_exp, len(headers))

    expenses = await models.get_expenses_range(db, start, end)
    for e in expenses:
        ws_exp.append([
            e["date"], e["amount"], e["category"],
            e.get("subcategory", ""), e.get("description", ""), e.get("source", "manual")
        ])
    for row in ws_exp.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    total_row = len(expenses) + 2
    ws_exp.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_exp.cell(row=total_row, column=2, value=sum(e["amount"] for e in expenses))
    ws_exp.cell(row=total_row, column=2).font = Font(bold=True)
    ws_exp.cell(row=total_row, column=2).number_format = CURRENCY_FORMAT
    _auto_width(ws_exp)

    # --- Sheet 2: Income ---
    ws_inc = wb.create_sheet("Income")
    headers = ["Date", "Amount", "Source", "Description"]
    ws_inc.append(headers)
    _style_header(ws_inc, len(headers))

    income = await models.get_income_range(db, start, end)
    for i in income:
        ws_inc.append([i["date"], i["amount"], i["source"], i.get("description", "")])
    for row in ws_inc.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT

    total_row = len(income) + 2
    ws_inc.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws_inc.cell(row=total_row, column=2, value=sum(i["amount"] for i in income))
    ws_inc.cell(row=total_row, column=2).font = Font(bold=True)
    ws_inc.cell(row=total_row, column=2).number_format = CURRENCY_FORMAT
    _auto_width(ws_inc)

    # --- Sheet 3: Food Log ---
    ws_food = wb.create_sheet("Food Log")
    headers = ["Date", "Meal", "Description", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    ws_food.append(headers)
    _style_header(ws_food, len(headers))

    food = await models.get_food_range(db, start, end)
    for f in food:
        ws_food.append([
            f["date"], f["meal_type"], f["description"],
            f.get("estimated_calories"), f.get("estimated_protein"),
            f.get("estimated_carbs"), f.get("estimated_fat"),
        ])
    _auto_width(ws_food)

    # --- Sheet 4: Health Data ---
    ws_health = wb.create_sheet("Health Data")
    headers = [
        "Date", "Sleep Score", "Sleep Hours", "Deep Sleep", "REM Sleep",
        "Steps", "Resting HR", "HRV", "SpO2", "Active Mins", "Calories Burned"
    ]
    ws_health.append(headers)
    _style_header(ws_health, len(headers))

    fitbit = await models.get_fitbit_range(db, start, end)
    for f in fitbit:
        ws_health.append([
            f["date"], f.get("sleep_score"), f.get("sleep_hours"),
            f.get("deep_sleep_mins"), f.get("rem_sleep_mins"), f.get("steps"),
            f.get("resting_hr"), f.get("hrv"), f.get("spo2"),
            f.get("active_zone_mins"), f.get("calories_burned"),
        ])
    _auto_width(ws_health)

    # --- Sheet 5: Monthly Summary ---
    ws_summary = wb.create_sheet("Monthly Summary")
    headers = ["Category", "Total Spent", "Transaction Count", "% of Total"]
    ws_summary.append(headers)
    _style_header(ws_summary, len(headers))

    categories = await models.get_expenses_by_category(db, start, end)
    grand_total = sum(c["total"] for c in categories) if categories else 1
    for c in categories:
        pct = c["total"] / grand_total * 100 if grand_total > 0 else 0
        ws_summary.append([c["category"], c["total"], c["count"], round(pct, 1)])
    for row in ws_summary.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = CURRENCY_FORMAT
    _auto_width(ws_summary)

    # Pie chart for categories
    if categories:
        pie = PieChart()
        pie.title = "Spending by Category"
        pie.width = 18
        pie.height = 12
        cat_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=len(categories) + 1)
        val_ref = Reference(ws_summary, min_col=2, min_row=2, max_row=len(categories) + 1)
        pie.add_data(val_ref, titles_from_data=False)
        pie.set_categories(cat_ref)
        ws_summary.add_chart(pie, "F2")

    # Line chart for daily spending trend on Expenses sheet
    if expenses:
        from collections import defaultdict
        daily_totals: dict[str, float] = defaultdict(float)
        for e in expenses:
            daily_totals[e["date"]] += e["amount"]

        ws_trend = wb.create_sheet("Daily Trend")
        ws_trend.append(["Date", "Amount"])
        _style_header(ws_trend, 2)
        for d_key in sorted(daily_totals):
            ws_trend.append([d_key, daily_totals[d_key]])
        for row in ws_trend.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = CURRENCY_FORMAT
        _auto_width(ws_trend)

        if len(daily_totals) > 1:
            line = LineChart()
            line.title = "Daily Spending Trend"
            line.y_axis.title = "Amount (₹)"
            line.width = 20
            line.height = 12
            cats = Reference(ws_trend, min_col=1, min_row=2, max_row=len(daily_totals) + 1)
            vals = Reference(ws_trend, min_col=2, min_row=1, max_row=len(daily_totals) + 1)
            line.add_data(vals, titles_from_data=True)
            line.set_categories(cats)
            ws_trend.add_chart(line, "D2")

    # Save
    Config.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"life_tracker_report_{date.today().isoformat()}.xlsx"
    filepath = Config.EXPORT_DIR / filename
    wb.save(str(filepath))
    logger.info("Excel report saved to %s", filepath)
    return filepath
